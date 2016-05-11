from tkinter import *
from tkinter import ttk
import tkinter.messagebox
from tabulate import tabulate
from PIL import Image, ImageTk
import openpyxl
import os
import pickle

###############################################################################################

# defining the functions


def editin():

    #opens excel file with system command
    
    os.system("start Book1.xlsx")
    lab5.configure(text="Editing Input", fg="red")


def generate():
    
    #wb points to the excel file
    wb = openpyxl.load_workbook("Book1.xlsx")
    f = open("input.txt", "w")
    fh = open("hours.txt","w")

    #sheet points to the sheet 1 in the input file
    
    sheet = wb.get_sheet_by_name('Sheet1')

    for i in range(3,25,2):
        for j in range(2, 500):
            var = sheet.cell(row=j, column=i).value
            #print(var)
            if var != None:
                f.write(sheet.cell(row=1, column=i).value)
                fh.write(sheet.cell(row=1, column=i).value)
                f.write('\n')
                fh.write('\n')
                f.write(sheet.cell(row=j, column=1).value)
                f.write('\n')
                f.write(sheet.cell(row=j, column=2).value)
                f.write('\n')

                if var != 'lab':
                    f.write(var)
                    f.write('\n')
                    fh.write(var)
                    fh.write('\n')
                    f.write('subject')
                    f.write('\n')
                else:
                    f.write('lab')
                    f.write('\n')
                    fh.write('lab')
                    fh.write('\n')
                fh.write(str(sheet.cell(row=j, column=i+1).value))
                fh.write('\n')

    f.write('end')
    fh.write('end')
    f.close()
    fh.close

    

    lab5.configure(text="Generation Complete", fg="green")

    #insert data about constraints to a text file. Data about selected and unselected ones
    
    f=open("constraints.txt","wb")
    coun=0
    lis=[]
    for i in range(5):
        var1 = cbx[i].get()
        if var1 == 1:
            coun+=1
        lis.append(var1)
    lis.insert(0,coun)
    print(lis)
    pickle.dump(lis,f)
    f.close()
    os.system("python johnylogic.py")

    f=open("johny.txt","rb")
    #ou=0
    #out=[]
    while(1):
        try:
            #out[ou] = Label(frame4,text="h")
            #out[ou].pack(side=TOP)
            ch=pickle.load(f)
            ch1=[ch,'','','','','','','']
            mon=pickle.load(f)
            mon.insert(0,'MON : ')
            tue=pickle.load(f)
            tue.insert(0,'TUE : ')
            wed=pickle.load(f)
            wed.insert(0,'WED : ')
            thu=pickle.load(f)
            thu.insert(0,'THU : ')
            fri=pickle.load(f)
            fri.insert(0,'FRI : ')
            main=[ch1,mon,tue,wed,thu,fri]
            print(tabulate(main))
            #out[ou].configure(text=tabulate(main))
            #ou+=1
        except:
            #print("br")
            break
    f.close()
    
    f=open("time.txt","rb")
    ch=pickle.load(f)
    print("Time required to complete : "+ch+" seconds")
    f.close()
    
    


def closer():
    if tkinter.messagebox.askyesno("Quit", "Are you sure you want to exit?"):
        root.destroy()


def login():

    #specifies the username and password required for accessing the program
    
    if entry1.get() == "root" and entry2.get() == "root":

        frame2.destroy()
        frame1.destroy()

        frame3.pack()
        frame4.pack(side=BOTTOM, fill=X)
        frame3.tkraise()
        frame4.tkraise()
    else:
        entry1.delete(0, END)
        entry2.delete(0, END)
        labstat.configure(text="Invalid", fg="red")


def setcons():
    frame3.pack_forget()
    frame4.pack_forget()
    frame5.pack()
    frame6.pack(side=BOTTOM, fill=X)
    frame5.tkraise()
    frame6.tkraise()


def con_done():
    frame5.pack_forget()
    frame6.pack_forget()
    frame3.pack()
    frame4.pack(side=BOTTOM, fill=X)
    frame3.tkraise()
    frame4.tkraise()

    lab5.configure(text="Constraints Specified",fg="green")

    edit.grid(sticky=N)
    setcon.grid(sticky=N)
    generat.grid(sticky=N)
    close1.pack(side=TOP)
    lab5.pack(fill=X)

def on_enter1(q):
    labstat3.configure(text="The HOD should not be assigned 4th periods on friday")
def on_leave1(q):
    labstat3.configure(text="")
def on_enter2(q):
    labstat3.configure(text="Only 1 library hour in a week")
def on_leave2(q):
    labstat3.configure(text="")
def on_enter3(q):
    labstat3.configure(text="Not more than 3 hours for a subject in a day")
def on_leave3(q):
    labstat3.configure(text="")
def on_enter4(q):
    labstat3.configure(text="Constraint_4")
def on_leave4(q):
    labstat3.configure(text="")
def on_enter5(q):
    labstat3.configure(text="Constraint_5")
def on_leave5(q):
    labstat3.configure(text="")

def closent(q):
    labstat.configure(text="Click to Exit",fg="red")
def closeleave(q):
    labstat.configure(text="Enter Username and Password to Login",fg="black")
def closent1(q):
    lab5.configure(text="Click to Exit", fg="red")
def closeleave1(q):
    lab5.configure(text="")

def ent1(q):
    lab5.configure(text="Edit the Input Excel File, to set the teachers and subjects", fg="blue")
def lea1(q):
    lab5.configure(text="")
def ent2(q):
    lab5.configure(text="To Generate the Time Table for given input teachers and selected Constraints", fg="blue")
def lea2(q):
    lab5.configure(text="")
def ent3(q):
    lab5.configure(text="To select the required constraints ", fg="blue")
def lea3(q):
    lab5.configure(text="")


#############################################################################

# defining window and frames
root = Tk()
root.title("Time Table Generator")
root.geometry("1280x720")

#login windows frames
frame1 = Frame(root)
frame1.pack()
frame2 = Frame(root)
frame2.pack(side="bottom", fill=X)

#main page frames
frame3 = Frame(root)
frame4 = Frame(root)

#constraint adding window frames
frame5 = Frame(root)
frame6 = Frame(root)


#############################################################################

# defining the entities

#login window entities
entry1 = Entry(frame1)

entry2 = Entry(frame1, show="\u2022")
lab1 = Label(frame1, text="USER", font="Calibri", fg="red", activebackground="yellow")
lab2 = Label(frame1, text="PASSWORD", font="Calibri", fg="red")
login = Button(frame1, text="LOGIN", font="Calibri", command=login)
close = Button(frame2, text="Close", font="Calibri", fg="red", command=closer)
photo = Image.open('timetable.jpg')
render = ImageTk.PhotoImage(photo)
labstat = Label(frame2, text="Enter Username and Password to Login", font="Calibri", anchor=W, relief=SUNKEN)
lab4 = Label(frame1, image=render)


#main window entities

close1 = Button(frame4, text="Close", font="Calibri", fg="red", command=closer)
lab5 = Label(frame4, text="Login Successfull", fg="green", font="Calibri", relief="sunken", anchor=W)

generat = Button(frame3, text="Generate", font="Calibri", command=generate)
edit = Button(frame3, text="Edit Input", font="Calibri", command=editin)
setcon = Button(frame3, text="Set Constraints", font="Calibri", command=setcons)


#constraint window entities

cbx=[IntVar(),IntVar(),IntVar(),IntVar(),IntVar()]


labstat3 = Label(frame6, text="Select Constraints to be added", font="Calibri", relief="sunken", anchor=W)
done = Button(frame6, text="Done", font="Calibri", fg="green", command=con_done)
cb1 = Checkbutton(frame5, text="Friday 4th period not for HOD",font="Calibri",variable=cbx[0],activeforeground="green")
cb2 = Checkbutton(frame5, text="Only 1 library hour in a week",font="Calibri",variable=cbx[1],activeforeground="green")
cb3 = Checkbutton(frame5, text="Not more than 3 hours for a subject in a day",font="Calibri",variable=cbx[2],activeforeground="green")
cb4 = Checkbutton(frame5, text="Constraint_4",font="Calibri",variable=cbx[3],activeforeground="green")
cb5 = Checkbutton(frame5, text="Constraint_5",font="Calibri",variable=cbx[4],activeforeground="green")

#binding actions of mouse on various entites (checkboxes)
cb1.bind("<Enter>", on_enter1)
cb1.bind("<Leave>", on_leave1)
cb2.bind("<Enter>", on_enter2)
cb2.bind("<Leave>", on_leave2)
cb3.bind("<Enter>", on_enter3)
cb3.bind("<Leave>", on_leave3)
cb4.bind("<Enter>", on_enter4)
cb4.bind("<Leave>", on_leave4)
cb5.bind("<Enter>", on_enter5)
cb5.bind("<Leave>", on_leave5)

#################################################################################

# packing/printing them
lab1.grid(row=15, column=15, sticky="E")
lab2.grid(row=16, column=15, sticky="E")
entry1.grid(row=15, column=16)
entry2.grid(row=16, column=16)
login.grid(row=17, column=15, columnspan=2)

lab4.grid(row=20)
labstat.pack(side=BOTTOM, fill=X)
close.pack(side="bottom")

edit.grid(sticky=N)
setcon.grid(sticky=N)
generat.grid(sticky = N)

close1.pack(side=TOP)
lab5.pack(fill=X)
edit.bind("<Enter>", ent1)
edit.bind("<Leave>", lea1)
generat.bind("<Enter>", ent2)
generat.bind("<Leave>", lea2)
setcon.bind("<Enter>", ent3)
setcon.bind("<Leave>", lea3)

close.bind("<Enter>", closent)
close.bind("<Leave>", closeleave)
close1.bind("<Enter>", closent1)
close1.bind("<Leave>", closeleave1)


cb1.pack()
#cb2.pack()
cb3.pack()
#cb4.pack()
#cb5.pack()

done.pack(side=TOP)
labstat3.pack(fill=X)

############################################################################

#specifying the main loop of the root window

root.protocol("WM_DELETE_WINDOW", closer)
root.mainloop()
